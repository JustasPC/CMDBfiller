from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import socket
import os

cmdbUrl="https://arscmdb.telia.se/arsys/forms/arscmdb/SHR%3ALandingConsole/Default+Administrator+View/

ip=[
    '//*[@id="arid_WIN_5_700000154"]',
    '//*[@id="arid_WIN_5_700000155"]',
    '//*[@id="arid_WIN_5_700000156"]',
    '//*[@id="arid_WIN_5_700000157"]',
    '//*[@id="arid_WIN_5_536871951"]',
    '//*[@id="arid_WIN_5_700001165"]',
    '//*[@id="arid_WIN_5_700001110"]'
    ]

ipv4=[
    '//*[@id="arid_WIN_4_700000154"]',
    '//*[@id="arid_WIN_4_700000155"]',
    '//*[@id="arid_WIN_4_700000156"]',
    '//*[@id="arid_WIN_4_700000157"]',
    '//*[@id="arid_WIN_4_536871951"]',
    '//*[@id="arid_WIN_4_700001165"]',
    '//*[@id="arid_WIN_4_700001110"]'
    ]

dnsv4=[
    '//*[@id="arid_WIN_4_700000153"]',
     '//*[@id="arid_WIN_4_700000101"]',
     '//*[@id="arid_WIN_4_700000102"]',
     '//*[@id="arid_WIN_4_700000103"]',
     '//*[@id="arid_WIN_4_700001166"]',
     '//*[@id="arid_WIN_4_700001167"]',
     '//*[@id="arid_WIN_4_700001109"]'
    ]

dns=[
    '//*[@id="arid_WIN_5_700000153"]',
     '//*[@id="arid_WIN_5_700000101"]',
     '//*[@id="arid_WIN_5_700000102"]',
     '//*[@id="arid_WIN_5_700000103"]',
     '//*[@id="arid_WIN_5_700001166"]',
     '//*[@id="arid_WIN_5_700001167"]',
     '//*[@id="arid_WIN_5_700001109"]'
    ]

subnet=[
        'arid_WIN_4_700000158',
        'arid_WIN_4_700000159',
        'arid_WIN_4_700000160',
        'arid_WIN_4_700000161',
        'arid_WIN_4_700001173',
        'arid_WIN_4_700001111'
    ]

gateway=[
         'arid_WIN_4_700000162',
         'arid_WIN_4_700000163',
         'arid_WIN_4_700000164',
         'arid_WIN_4_700000165',
         'arid_WIN_4_700001180',
         'arid_WIN_4_700001181',
         'arid_WIN_4_1000005996'
         ]

dnsres=[
    'arid_WIN_4_700001155',
    'arid_WIN_4_700001159'
    ]

vlan=[
    'arid_WIN_4_700000142',
    'arid_WIN_4_700000143',
    'arid_WIN_4_1000006010'
    ]

fire=[
    'arid_WIN_4_700000138'
    ]

network=[
    'arid_WIN_4_700000121',
    'arid_WIN_4_700000122'
    ]

level=[
    'arid_WIN_4_700000125',
    'arid_WIN_4_700000126'
    ]

ciname=[
    'arid_WIN_4_200000020'
    ]

cellNames=[
    'New_Done'
    'Order_nr',
    'Prod_BMN_Ilo',
    'CI',
    'IP4',
    'Subnet',
    'Gateway',
    'DNS',
    'DNSresolver1',
    'DNSresolver2',
    'Network',
    'NetSecLvl',
    'VLANid',
    'Fire'
    ]
net_tabs = {
    "Network":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[6]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/div[2]/div[2]/div/dl/dd[12]/span[2]/a",
    "Network1":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[6]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[1]/span[2]/a",
    "Network2":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[6]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[2]/span[2]/a",
    "Network3":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[6]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[3]/span[2]/a",
    "Network4":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[6]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[4]/span[2]/a",
    "Network5":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[6]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[5]/span[2]/a",
    "Network6":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[6]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[6]/span[2]/a",
    "Console":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[6]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[7]/span[2]/a"
    }
net_tabsv4 = {
    "Network":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/div[2]/div[2]/div/dl/dd[12]/span[2]/a",
    "Network1":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[1]/span[2]/a",
    "Network2":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[2]/span[2]/a",
    "Network3":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[3]/span[2]/a",
    "Network4":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[4]/span[2]/a",
    "Network5":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[5]/span[2]/a",
    "Network6":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[6]/span[2]/a",
    "Console":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[7]/span[2]/a"
    }

cmdb_gui = {
    "first_ci":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[2]/div/div[2]/div/div[2]/table/tbody/tr[2]",
    "ci_start":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[2]/div/div[2]/div/div[2]/table/tbody/tr[",
    "ci_end":"]/td[1]/nobr/span",
    "result_cis":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[2]/div/div[2]/div/div[2]/table/tbody/tr[3]/td[1]/nobr/span",
    "second_ci":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[2]/div/div[2]/div/div[2]/table/tbody/tr[keiciama vieta]",
    "res_numb":"/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[2]/div/div[1]/table/tbody/tr/td[2]",
    }


browser = webdriver.Firefox()
excelis = openpyxl.Workbook()
url = 'https://arscmdb.telia.se/arsys/forms/arscmdb/SHR%3ALandingConsole/Default+Administrator+View/

##########################Web tikrinimas
def fun_web_check():
    ats = False
    try:
        elem = browser.find_element_by_id(ciname[0])
        ats = True
    except:
        print('Failed web check')
        ats = False
    return ats


#######################Darbo pradzia su exceliu ir jo kurimas
def fun_load_excel():
    global excelis
    control = False
    try:
        excelis = openpyxl.load_workbook('CMDBimport.xlsx')
        control=True
    except:
        print('Failed excel check')
        new_excelis()
    return control

def fun_new_excelis():
    print('Bus sukurtas naujas CMDBimport.xlsx')
    try:
        excelis = openpyxl.load_workbook('CMDBimport.xlsx')
        print('Failas jau egzistuoja')
        print('Perrašyti egzistuojanti?')
        option = input('y/n?: ')
        if(option == 'Y' or option == 'y' or option == 'yes'):
            fun_newWork()
        else:
            print('Pasirinkta nekurti naujo')
    except:
        fun_newFile()

def fun_newFile():    
    excel = openpyxl.Workbook()
    excel['Sheet'].title='IPtoCMDB'
    Sheet=excel['IPtoCMDB']
    for name in cellNames:
        Sheet.cell(row=1, column=(cellNames.index(name)+1)).value=name
    excel.save('CMDBimport.xlsx')
    print('Naujas exelis sukurtas')

##########################LogName
def fun_get_log_name():
    line = time.asctime()
    textas = line.split(' ')
    line = textas[4].split(':')
    try:
        tmp = 'CMDBLog'+textas[5]+textas[1]+textas[3]+line[0]+line[1]+line[2]+'.txt'
    except:
        line = textas[3].split(':')
        tmp = 'CMDBLog'+textas[4]+textas[1]+textas[2]+line[0]+line[1]+line[2]+'.txt'
    return tmp
##########################

def fun_pildymas():
    elem = browser.find_element_by_id(ip[0])
    if(not(elem.is_displayed())):
        elem = find_element_by_xpath('/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/div[2]/div[2]/div/dl/dd[12]/span[2]/a')
        elem.click()

def fun_prod():
    global excelis
    elem = browser.find_element_by_xpath('/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[4]/fieldset/div/div/div/div[4]/div[6]/div/div/div[2]/fieldset/div/div[21]/fieldset[12]/div[7]/div[2]/div[2]/div/dl/dd[1]/span[2]/a')
    elem.click()
    sheet = exelis['IPtoCMDB']
    

def fun_login_screen():
    browser.get(cmdbUrl)
    try:
        alert = browser.switch_to.alert()
        alert.send_keys(Keys.ENTER)
    except:
        print('Loading...')
    print("Login and navigate to Asset Management Console")


def fun_collect_ips_new_excel():
    excelis = openpyxl.Workbook()
    excelis.save('collectIPs.xlsx')
    excelis.save('collectIPsResults.xlsx')



def fun_collect_ips_fill_excel():
    excelis = openpyxl.load_workbook('collectIPs.xlsx')
    sheet = excelis['Sheet']
    excelis2 = openpyxl.load_workbook('collectIPsResults.xlsx')
    res_sheet = excelis2['Sheet']
    cache={}
    num = sheet.max_row
    elem = browser.find_element_by_xpath(cmdb_gui["res_numb"])
    tmp = elem.text.split()
    if(num == int(tmp[0])):
        print('INFO:****Same number of elements*****')
    else:
        print('INFO:****Not the same number of elements*****')
    for i in range(0,num):
        xpath = cmdb_gui["ci_start"]+str(2+i)+cmdb_gui["ci_end"]
        elem = browser.find_element_by_xpath(xpath)
        cache.update({elem.text : xpath})
    print(cache)
    for i in range(1,num+1):
        value=sheet.cell(row=i,column=3).value.lower().strip()
        try:
            elem = browser.find_element_by_xpath(cache[value])
        except:
            elem = browser.find_element_by_xpath(cache[str(value).upper()])
        print(value)
        elem.click()
        time.sleep(1)
        counter=0
        for tab, xpath in net_tabsv4.items():
            if counter==0:
                elem = browser.find_element_by_xpath(xpath)
                elem.click()
            else:
                elem = browser.find_element_by_xpath(xpath)
                elem.click()
                time.sleep(1)
                elem = browser.find_element_by_xpath(ip[counter-1])
                elem2 = browser.find_element_by_xpath(dns[counter-1])
                if (len(elem.get_attribute('value'))>4 or len(elem2.get_attribute('value'))>4):
                    eile=res_sheet.max_row
                    res_sheet.cell(row=eile+1,column=1).value=sheet.cell(row=i,column=1).value
                    print(res_sheet.cell(row=eile+1,column=1).value)
                    res_sheet.cell(row=eile+1,column=2).value=sheet.cell(row=i,column=2).value
                    print(res_sheet.cell(row=eile+1,column=2).value)
                    res_sheet.cell(row=eile+1,column=3).value=value
                    res_sheet.cell(row=eile+1,column=4).value=elem.get_attribute('value')
                    print(res_sheet.cell(row=eile+1,column=4).value)
                    res_sheet.cell(row=eile+1,column=5).value=elem2.get_attribute('value')
                    print(res_sheet.cell(row=eile+1,column=5).value)
                    try:
                        name = socket.gethostbyaddr(elem.get_attribute('value'))
                        if(not(name[0]==elem2.get_attribute('value'))):
                            res_sheet.cell(row=eile+1,column=6).value=name[0]
                    except:
                        print('Nslookup crash')
            counter=counter+1
    excelis2.save('collectIPsResults.xlsx')
